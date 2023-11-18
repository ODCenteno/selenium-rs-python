import datetime
import openpyxl

# Ubicando el libro donde tendemos la información
book = openpyxl.load_workbook("path/archive.xlsx")

# Seleccionando la información celda por celda
sheet = book.active
cell = sheet.cell(row=1, column=2)
print(cell.value)

# Data can be assigned directly to cells
sheet['A1'] = 42

# Rows can also be appended
sheet.append([1, 2, 3])

# Python types will automatically be converted
sheet['A2'] = datetime.datetime.now()

# Save the file
sheet.save("sample.xlsx")


# Iterando a través de toda la tabla por columnas
for j in range(1, sheet.max_column + 1):
    for i in range(1, sheet.max_row + 1):
        print(sheet.cell(row=i, column=j).value)

# Iterando a través de toda la tabla por filas
for i in range(1, sheet.max_row + 1):
    for j in range(1, sheet.max_column + 1):
        print(sheet.cell(row=i, column=j).value)

# Construyendo un diccionario desde con los datos de excel
# Se itera a través de las filas
new_dict = {}
for i in range(1, sheet.max_row + 1):
    # Se puede elegir una sola fila con if:
    if sheet.cell(row=1, column=1).value == "Testcase 2":
        for j in range(1, sheet.max_column + 1):
            new_dict[sheet.cell(row=1, column=j).value] = sheet.cell(
                row=i, column=j).value

# Generando un método estático para utilizarlo en clases


@staticmethod
def get_test_data(test_name):
    new_dict = {}

    book = openpyxl.load_workbook("path/archive.xlsx")

# Seleccionando la información celda por celda
    sheet = book.active

    for i in range(1, sheet.max_row + 1):
        # Se puede elegir una sola fila con if:
        if sheet.cell(row=1, column=1).value == test_name:
            for j in range(2, sheet.max_column + 1):
                new_dict[sheet.cell(row=1, column=j).value] = sheet.cell(
                    row=i, column=j).value

    return [dict]
